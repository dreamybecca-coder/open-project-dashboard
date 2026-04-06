#!/usr/bin/env python3
"""
将飞书多维表格数据导出为Excel
"""

import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 读取原始数据
with open("records_raw.json", "r", encoding="utf-8") as f:
    records = json.load(f)

print(f"Total records: {len(records)}")

# 定义字段名称映射（按顺序）
fields = [
    ("案主编号", "case_owner_id"),
    ("姓名", "name"),
    ("微信昵称", "wechat_nickname"),
    ("所在城市及联系方式", "location_contact"),
    ("所在大区", "region"),
    ("案主个人介绍", "self_intro"),
    ("学分", "credit_score"),
    ("项目课题介绍", "project_intro"),
    ("航海目标和计划", "sailing_goal_plan"),
    ("项目想突破的方向", "breakthrough_direction"),
    ("案主报名形式", "registration_type"),
    ("案主是否选择'纯MBA战队'模式", "mba_team_mode"),
    ("案主是否选择'开源挑战'模式", "open_source_mode"),
    ("船票（押券数）", "ticket_count"),
    ("一页纸完整介绍", "one_page_intro"),
    ("一堂学习故事", "learning_story"),
]

# 提取数据
data_rows = []
for record in records:
    fields_data = record.get("fields", {})
    row = {}
    for cn_name, en_name in fields:
        value = fields_data.get(cn_name, "")
        # 处理多选字段（转为逗号分隔字符串）
        if isinstance(value, list):
            value = ", ".join(str(v) for v in value)
        row[en_name] = value
    data_rows.append(row)

# 创建DataFrame
df = pd.DataFrame(data_rows)

# 保存为Excel
output_file = "开源案主清单.xlsx"
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='案主清单', index=False)
    
    # 获取工作表
    ws = writer.sheets['案主清单']
    
    # 设置表头样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置数据区域样式
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(vertical='top', wrap_text=True)
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # 设置列宽
    column_widths = {
        'A': 12,   # 案主编号
        'B': 10,   # 姓名
        'C': 15,   # 微信昵称
        'D': 30,   # 所在城市及联系方式
        'E': 12,   # 所在大区
        'F': 50,   # 案主个人介绍
        'G': 8,    # 学分
        'H': 50,   # 项目课题介绍
        'I': 50,   # 航海目标和计划
        'J': 20,   # 项目想突破的方向
        'K': 18,   # 案主报名形式
        'L': 18,   # 纯MBA战队模式
        'M': 20,   # 开源挑战模式
        'N': 10,   # 船票数
        'O': 50,   # 一页纸完整介绍
        'P': 100,  # 一堂学习故事
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 设置行高
    ws.row_dimensions[1].height = 30

print(f"Excel file saved to: {output_file}")
print(f"Total rows: {len(df)}")
